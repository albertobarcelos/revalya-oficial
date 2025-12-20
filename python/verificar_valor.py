import openpyxl

# Carrega a planilha
workbook = openpyxl.load_workbook('contratos_prontos_with_ids.xlsx')
sheet = workbook.active

# Verifica a linha 10, coluna 14 (PDV/Comandas)
row_num = 10
col_num = 14
value = sheet.cell(row=row_num, column=col_num).value

print(f'🔍 Valor na linha {row_num}, coluna {col_num} (PDV/Comandas):')
print(f'Valor: {value}')
print(f'Tipo: {type(value)}')

# Testa as condições do script
value_str = str(value).strip().upper() if value else ''
print(f'\n🧪 Testando condições do script:')
print(f'É "SIM", "1", "YES", "TRUE": {value_str in ["SIM", "1", "YES", "TRUE"]}')
print(f'É número > 0: {isinstance(value, (int, float)) and value > 0}')
print(f'É "NÃO", "NAO", "0", "NO", "FALSE": {value_str in ["NÃO", "NAO", "0", "NO", "FALSE"]}')

# Verifica também o valor da linha 10, coluna 23 (Balança Auto Serviço)
row_num = 10
col_num = 23
value = sheet.cell(row=row_num, column=col_num).value
print(f'\n🔍 Valor na linha {row_num}, coluna {col_num} (Balança Auto Serviço):')
print(f'Valor: {value}')
print(f'Tipo: {type(value)}')

value_str = str(value).strip().upper() if value else ''
print(f'\n🧪 Testando condições do script:')
print(f'É "SIM", "1", "YES", "TRUE": {value_str in ["SIM", "1", "YES", "TRUE"]}')
print(f'É número > 0: {isinstance(value, (int, float)) and value > 0}')

workbook.close()