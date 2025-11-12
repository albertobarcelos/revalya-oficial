#!/usr/bin/env python3
"""
Script para buscar contract_id pelo contract_number e preencher na planilha
"""

import openpyxl
import os
from dotenv import load_dotenv
from supabase import create_client, Client

# Carregar variáveis de ambiente
load_dotenv()

# Configurações do Supabase
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_KEY') or os.getenv('SUPABASE_KEY')
TENANT_ID = os.getenv('TENANT_ID', 'c9ed4c60-0b15-4d21-9c99-1d55e5b3e5f0')

def get_contract_id_by_number(supabase: Client, contract_number: str, tenant_id: str):
    """Busca o ID do contrato pelo número"""
    try:
        response = supabase.table('contracts').select('id').eq('contract_number', contract_number).eq('tenant_id', tenant_id).execute()
        if response.data and len(response.data) > 0:
            return response.data[0]['id']
        return None
    except Exception as e:
        print(f"Erro ao buscar contrato {contract_number}: {e}")
        return None

def main():
    """Função principal"""
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("Erro: SUPABASE_URL e SUPABASE_KEY são obrigatórios")
        return
    
    # Inicializar Supabase
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    # Carregar Excel
    excel_path = os.getenv('CONTRATOS_XLSX', 'contratos_prontos.xlsx')
    if not os.path.exists(excel_path):
        print(f"Erro: Arquivo {excel_path} não encontrado")
        return
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Identificar colunas
    contract_number_col = None
    contract_id_col = None
    
    # Buscar coluna do contract_number (CodGE)
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip().lower() in ['codge', 'cod_ge', 'contract_number']:
            contract_number_col = col
            break
    
    # Buscar ou criar coluna para contract_id
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and str(header).strip().lower() == 'contract_id':
            contract_id_col = col
            break
    
    # Se não encontrou coluna contract_id, criar nova coluna
    if not contract_id_col:
        contract_id_col = ws.max_column + 1
        ws.cell(row=1, column=contract_id_col, value='contract_id')
        print(f"Coluna 'contract_id' criada na posição {contract_id_col}")
    
    if not contract_number_col:
        print("Erro: Coluna 'CodGE' ou 'contract_number' não encontrada")
        return
    
    print(f"Processando contratos da coluna {contract_number_col} para coluna {contract_id_col}")
    
    # Estatísticas
    total_processados = 0
    encontrados = 0
    nao_encontrados = 0
    
    # Processar cada linha (começando da linha 3, pulando cabeçalhos)
    for row in range(3, ws.max_row + 1):
        contract_number = ws.cell(row=row, column=contract_number_col).value
        
        if not contract_number:
            continue
            
        total_processados += 1
        
        # Buscar ID do contrato
        contract_id = get_contract_id_by_number(supabase, str(contract_number), TENANT_ID)
        
        if contract_id:
            # Preencher contract_id na planilha
            ws.cell(row=row, column=contract_id_col, value=contract_id)
            encontrados += 1
            print(f"✓ Linha {row}: Contrato {contract_number} -> ID: {contract_id}")
        else:
            # Limpar valor existente se contrato não foi encontrado
            ws.cell(row=row, column=contract_id_col, value=None)
            nao_encontrados += 1
            print(f"✗ Linha {row}: Contrato {contract_number} não encontrado no banco")
    
    # Salvar planilha
    output_path = excel_path.replace('.xlsx', '_with_ids.xlsx')
    wb.save(output_path)
    
    # Relatório final
    print("\n" + "="*60)
    print("RELATÓRIO DE PREENCHIMENTO DE CONTRACT_IDS")
    print("="*60)
    print(f"Total de contratos processados: {total_processados}")
    print(f"Contratos encontrados: {encontrados}")
    print(f"Contratos não encontrados: {nao_encontrados}")
    print(f"Planilha salva em: {output_path}")

if __name__ == "__main__":
    main()