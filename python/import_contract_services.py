#!/usr/bin/env python3
"""
Script para importar serviços dos contratos do Excel para Supabase
"""

import openpyxl
import os
from datetime import datetime
from dotenv import load_dotenv
from supabase import create_client, Client

# Carregar variáveis de ambiente
load_dotenv()

# Configurações do Supabase
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_KEY') or os.getenv('SUPABASE_KEY')
TENANT_ID = '8d2888f1-64a5-445f-84f5-2614d5160251'  # Tenant correto baseado nos dados

# Mapeamento de serviços (UUIDs reais)
SERVICES_MAPPING = {
    'Gestao': 'dbad5192-79b1-41e6-adbd-5218167c738c',
    'PDV/Comandas': 'c1552361-c1db-43ae-ad3a-9a6f8143f668',
    'NFCE': 'c8cb99e1-3cea-4a99-ae93-5de95d45e39f',
    'Estoque': '86f31600-69f9-4426-82f4-ff9f92c54021',
    'Financeiro': '2c343d48-fea9-4002-9106-4a324b5a5189',
    'Delivery Legal': '8b009d88-9219-4e97-90a2-8bb3677b8ec7',
    'Delivery Legal +': '6f215517-b962-4544-a6b9-111555809e14',
    'Fidelidade legal': '3e64cc33-9948-47d2-8d13-9cdb757c8bf1',
    'Totem Autoatendimento': '565022c0-a844-412a-a81b-e41529c513c3',
    'KDS': None,  # Precisa ser descoberto
    'Balanca Auto Servico': None  # Precisa ser descoberto
}

def get_contract_id_by_number(supabase: Client, contract_number: str, tenant_id: str):
    """Busca o ID do contrato pelo número"""
    try:
        response = supabase.table('contracts').select('id').eq('contract_number', contract_number).eq('tenant_id', tenant_id).execute()
        if response.data and len(response.data) > 0:
            return response.data[0]['id']
        return None
    except Exception as e:
        print(f"Erro ao buscar contrato {contract_number}: {e}")
        return None

def parse_service_value(value):
    """Converte valor do serviço para quantidade"""
    if not value or str(value).strip().upper() in ['NAO', 'NÃO', 'NO', 'FALSE', '0']:
        return 0
    
    if str(value).strip().upper() in ['SIM', 'YES', 'TRUE', '1']:
        return 1
    
    # Se for número, retorna o número
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return 0

def create_contract_service(supabase: Client, contract_id: str, service_id: str, quantity: int, tenant_id: str, cost_price: float = 0.0):
    """Cria um serviço de contrato no Supabase"""
    try:
        contract_service_data = {
            'contract_id': contract_id,
            'service_id': service_id,
            'quantity': quantity,
            'unit_price': 35.00,  # Preço padrão por unidade
            'discount_percentage': 0,
            'description': 'Serviço importado da planilha',
            'is_active': True,
            'tenant_id': tenant_id,
            'payment_method': 'Boleto',
            'billing_type': 'Único',
            'recurrence_frequency': 'Mensal',
            'installments': 1,
            'due_type': 'days_after_billing',
            'due_value': 5,
            'generate_billing': False,  # FALSE conforme schema
            'cost_price': cost_price,  # Custo do serviço
            'created_at': datetime.now().isoformat(),
            'updated_at': datetime.now().isoformat()
        }
        
        response = supabase.table('contract_services').insert(contract_service_data).execute()
        return True, None
        
    except Exception as e:
        return False, str(e)

def main():
    """Função principal"""
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("Erro: SUPABASE_URL e SUPABASE_KEY são obrigatórios")
        return
    
    # Inicializar Supabase
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    # Carregar Excel
    excel_path = 'python/contratos_prontos_with_ids.xlsx'
    if not os.path.exists(excel_path):
        print(f"Erro: Arquivo {excel_path} não encontrado")
        return
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    # Identificar colunas de serviços
    service_columns = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header and header.strip() in SERVICES_MAPPING:
            service_columns[header.strip()] = col
    
    print(f"Serviços encontrados na planilha: {list(service_columns.keys())}")
    
    # Estatísticas
    total_linhas = 0
    servicos_criados = 0
    servicos_ignorados = 0
    erros = []
    
    # Processar cada linha (começando da linha 3, pulando cabeçalhos)
    for row in range(3, ws.max_row + 1):
        total_linhas += 1
        
        # Obter dados do contrato
        contract_number = ws.cell(row=row, column=2).value  # CodGE
        customer_id = ws.cell(row=row, column=3).value
        cost_value = ws.cell(row=row, column=10).value  # Custo
        contract_id = ws.cell(row=row, column=23).value  # contract_id
        
        if not contract_number or not customer_id or not contract_id:
            print(f"Linha {row}: Ignorada - contrato, cliente ou contract_id vazio")
            continue
        
        # Calcular custo por serviço (dividir custo total pelos serviços ativos)
        active_services = 0
        for service_name, col_idx in service_columns.items():
            value = ws.cell(row=row, column=col_idx).value
            quantity = parse_service_value(value)
            if quantity > 0:
                active_services += 1
        
        cost_per_service = 0.0
        if active_services > 0 and cost_value:
            try:
                cost_per_service = float(cost_value) / active_services
            except (ValueError, TypeError):
                cost_per_service = 0.0
        
        # Processar cada serviço
        for service_name, col_idx in service_columns.items():
            value = ws.cell(row=row, column=col_idx).value
            quantity = parse_service_value(value)
            
            if quantity == 0:
                servicos_ignorados += 1
                continue
            
            service_id = SERVICES_MAPPING.get(service_name)
            if not service_id:
                print(f"Aviso: ID do serviço '{service_name}' não encontrado")
                continue
            
            # Criar serviço do contrato
            success, error = create_contract_service(
                supabase, 
                contract_id, 
                service_id, 
                quantity,
                TENANT_ID,
                cost_per_service if service_name == 'Gestao' else 0.0
            )
            
            if success:
                servicos_criados += 1
                print(f"✓ Contrato {contract_number}: Serviço {service_name} criado (Qtd: {quantity})")
            else:
                erros.append(f"Contrato {contract_number}, Serviço {service_name}: {error}")
                print(f"✗ Contrato {contract_number}: Erro ao criar serviço {service_name}: {error}")
    
    # Relatório final
    print("\n" + "="*60)
    print("RELATÓRIO DE IMPORTAÇÃO DE SERVIÇOS")
    print("="*60)
    print(f"Total de linhas processadas: {total_linhas}")
    print(f"Serviços criados: {servicos_criados}")
    print(f"Serviços ignorados (quantidade 0): {servicos_ignorados}")
    print(f"Erros: {len(erros)}")
    
    if erros:
        print("\nErros detalhados:")
        for erro in erros[:10]:  # Mostrar primeiros 10 erros
            print(f"  - {erro}")
        if len(erros) > 10:
            print(f"  ... e mais {len(erros) - 10} erros")

if __name__ == "__main__":
    main()