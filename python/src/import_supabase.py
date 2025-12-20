from typing import List, Dict, Any, Tuple
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from .schema import contrato_columns


def _parse_date(value) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    # tentar converter de string dd/mm/yyyy
    try:
        return datetime.strptime(str(value), "%d/%m/%Y").date().isoformat()
    except Exception:
        # tentar ISO diretamente
        try:
            return datetime.fromisoformat(str(value)).date().isoformat()
        except Exception:
            return str(value)


def _parse_number(value) -> Any:
    if value is None or value == "":
        return None
    try:
        # troca vírgula por ponto se vier em formato PT-BR
        return float(str(value).replace(".", "").replace(",", "."))
    except Exception:
        return value


def ler_planilha(path: str) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Lê a planilha e retorna (registros, erros).
    Considera a primeira aba como fonte de dados.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {path}")

    wb = load_workbook(p, data_only=True)
    ws = wb.active

    cols = contrato_columns()
    headers = [c["title"] for c in cols]
    name_map = {c["title"]: c["name"] for c in cols}
    type_map = {c["name"]: c.get("type", "text")}

    # Verificar cabeçalhos
    for idx, h in enumerate(headers, start=1):
        cell_val = ws.cell(row=1, column=idx).value
        if cell_val != h:
            raise ValueError(
                f"Cabeçalho inesperado na coluna {idx}: esperado '{h}', encontrado '{cell_val}'"
            )

    registros: List[Dict[str, Any]] = []
    erros: List[str] = []

    row = 2
    while True:
        # Parar quando a linha estiver toda vazia
        row_values = [ws.cell(row=row, column=i).value for i in range(1, len(cols) + 1)]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_values):
            break

        registro: Dict[str, Any] = {}
        for idx, col in enumerate(cols, start=1):
            val = ws.cell(row=row, column=idx).value
            key = col["name"]
            tipo = col.get("type", "text")

            if tipo == "date":
                registro[key] = _parse_date(val)
            elif tipo == "number":
                registro[key] = _parse_number(val)
            else:
                registro[key] = val if val is not None else None

            if col.get("required") and (registro[key] is None or str(registro[key]).strip() == ""):
                erros.append(f"Linha {row}: coluna '{col['title']}' é obrigatória.")

            if tipo == "list" and col.get("list_values") and registro[key] is not None:
                if str(registro[key]).lower() not in [v.lower() for v in col["list_values"]]:
                    erros.append(
                        f"Linha {row}: valor '{registro[key]}' inválido para '{col['title']}'. "
                        f"Permitidos: {', '.join(col['list_values'])}"
                    )

        registros.append(registro)
        row += 1

    return registros, erros


def importar_para_supabase(client, registros: List[Dict[str, Any]], tabela: str = "contratos") -> Any:
    """
    Envia os registros para a tabela indicada no Supabase.
    Por padrão usa insert; se o cliente suportar upsert, tentamos upsert.
    """
    table_api = client.table(tabela)

    # Tentar upsert se disponível
    try:
        result = table_api.upsert(registros).execute()
    except Exception:
        result = table_api.insert(registros).execute()

    return result