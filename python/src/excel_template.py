from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Optional

from .schema import contrato_columns


def gerar_template(output_path: Optional[str] = None) -> Path:
    """
    Gera a planilha Excel com cabeçalhos, formatos, filtros e validações.
    Retorna o caminho do arquivo gerado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Contratos"

    cols = contrato_columns()

    # Estilos de cabeçalho
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(border_style="thin", color="999999")
    header_border = Border(top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # Escrever cabeçalhos e configurar colunas
    for idx, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=idx, value=col["title"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = header_border
        ws.column_dimensions[get_column_letter(idx)].width = col.get("width", 15)

    # Congelar cabeçalho e habilitar autofiltro
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Formatação e validações por coluna
    max_rows = 5000  # limite razoável para validação/aplicação
    for idx, col in enumerate(cols, start=1):
        col_letter = get_column_letter(idx)

        # Formato numérico / data
        num_fmt = col.get("number_format")
        if num_fmt:
            for r in range(2, max_rows + 1):
                ws[f"{col_letter}{r}"].number_format = num_fmt

        # Validação de lista
        if col.get("type") == "list" and col.get("list_values"):
            lista = ",".join(col["list_values"])  # ex: "ativo,inativo,suspenso"
            dv = DataValidation(type="list", formula1=f'"{lista}"', allow_blank=not col.get("required", False))
            dv.prompt = f"Selecione um valor para {col['title']}"
            dv.error = "Valor inválido"
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{max_rows}")

        # Validação de data (opcional — Excel já formata)
        if col.get("type") == "date":
            # Não aplicamos DataValidation de data para evitar bloqueios — apenas formato.
            pass

    # Aba dicionário de dados
    dict_ws = wb.create_sheet("Dicionario")
    dict_ws.append(["Coluna", "Obrigatório", "Descrição", "Tipo", "Valores (se lista)"])
    for c in contrato_columns():
        dict_ws.append([
            c["title"],
            "Sim" if c.get("required") else "Não",
            c.get("description", ""),
            c.get("type", "text"),
            ", ".join(c.get("list_values", [])),
        ])

    # Caminho de saída
    out = Path(output_path or "contratos_template.xlsx")
    wb.save(out)
    return out