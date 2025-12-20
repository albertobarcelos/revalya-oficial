import os
import re
import csv
import unicodedata
from typing import Optional
from openpyxl import load_workbook

# Caminhos base
BASE_DIR = r'd:\DESENVOLVIMENTO\revalya-oficial\python'
INPUT_XLSX = os.path.join(BASE_DIR, 'contratos.xlsx')
OUTPUT_XLSX = os.path.join(BASE_DIR, 'contratos_atualizado_final.xlsx')
UNMATCHED_CSV = os.path.join(BASE_DIR, 'contratos_unmatched_final.csv')

def normalize_name(name: str) -> str:
    """Normaliza nome removendo acentos e espaços para comparar nomes de planilhas."""
    if not isinstance(name, str):
        name = str(name)
    nf = unicodedata.normalize('NFD', name)
    s = ''.join(c for c in nf if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', '', s).lower()

def find_sheet(wb, target: str):
    """Encontra a planilha pelo nome alvo, ignorando acentos e espaços."""
    target_norm = normalize_name(target)
    for nm in wb.sheetnames:
        if normalize_name(nm) == target_norm:
            return wb[nm]
    return None

DIGITS_RE = re.compile(r'\D+')

def cnpj_digits(value: Optional[object]) -> str:
    """Extrai apenas dígitos do valor de CNPJ, preservando zeros à esquerda."""
    if value is None:
        return ''
    if isinstance(value, int):
        s = str(value)
    elif isinstance(value, float):
        s = str(int(value)) if value.is_integer() else str(value)
    else:
        s = str(value)
    s = s.strip()
    return DIGITS_RE.sub('', s)

def parse_cost(val: Optional[object]):
    """Converte valor de custo para float quando possível, aceitando vírgula decimal."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return None
    s2 = re.sub(r'[^0-9.,-]', '', s)
    if ',' in s2:
        s2 = s2.replace('.', '')
        s2 = s2.replace(',', '.')
    try:
        return float(s2)
    except Exception:
        return s

def main():
    print('Abrindo arquivo:', INPUT_XLSX)
    wb = load_workbook(INPUT_XLSX, data_only=True)

    lic_sheet = find_sheet(wb, 'LicencasAtivas')
    pag_sheet = find_sheet(wb, 'Pagina1')
    if lic_sheet is None and 'LicencasAtivas' in wb.sheetnames:
        lic_sheet = wb['LicencasAtivas']
    if pag_sheet is None and 'Página1' in wb.sheetnames:
        pag_sheet = wb['Página1']
    if lic_sheet is None or pag_sheet is None:
        print('Erro: não foi possível localizar as planilhas necessárias.')
        print('Disponíveis:', wb.sheetnames)
        return

    # Mapa CNPJ -> custo a partir de Página1 (CNPJ na coluna C, custo na coluna E)
    mapa = {}
    linhas_pag = 0
    for r in range(2, pag_sheet.max_row + 1):
        cnpj_raw = pag_sheet.cell(row=r, column=3).value  # Coluna C
        custo_raw = pag_sheet.cell(row=r, column=5).value  # Coluna E
        cnpj_key = cnpj_digits(cnpj_raw)
        if cnpj_key:
            mapa[cnpj_key] = parse_cost(custo_raw)
            linhas_pag += 1
    print(f'Mapa criado a partir de Página1: {linhas_pag} linhas, {len(mapa)} chaves únicas')

    # Detecta coluna do CNPJ em LicencasAtivas
    cnpj_col = None
    header_row = 1
    for c in range(1, lic_sheet.max_column + 1):
        hv = lic_sheet.cell(row=header_row, column=c).value
        if hv and isinstance(hv, str) and 'cnpj' in hv.lower():
            cnpj_col = c
            break
    if cnpj_col is None:
        cnpj_col = 1  # fallback para coluna A
    print(f'Coluna CNPJ em LicencasAtivas: {cnpj_col}')

    # Destino: coluna H (8) com header "Custo"
    destino_col = 8
    lic_sheet.cell(row=1, column=destino_col).value = 'Custo'

    atualizados = 0
    sem_match = 0
    exemplos_unmatched = []
    unmatched_rows = []

    for r in range(2, lic_sheet.max_row + 1):
        raw = lic_sheet.cell(row=r, column=cnpj_col).value
        key = cnpj_digits(raw)
        if not key:
            sem_match += 1
            unmatched_rows.append({
                'row': r,
                'cnpj_raw': '' if raw is None else str(raw),
                'cnpj_digits': key,
                'reason': 'CNPJ vazio'
            })
            if len(exemplos_unmatched) < 10:
                exemplos_unmatched.append((r, raw, key, 'CNPJ vazio'))
            continue
        if key in mapa:
            lic_sheet.cell(row=r, column=destino_col).value = mapa[key]
            atualizados += 1
        else:
            sem_match += 1
            unmatched_rows.append({
                'row': r,
                'cnpj_raw': '' if raw is None else str(raw),
                'cnpj_digits': key,
                'reason': 'sem correspondência'
            })
            if len(exemplos_unmatched) < 10:
                exemplos_unmatched.append((r, raw, key, 'sem correspondência'))

    print(f'Salvando arquivo atualizado em: {OUTPUT_XLSX}')
    wb.save(OUTPUT_XLSX)

    print(f'Gerando relatório de não casados em: {UNMATCHED_CSV}')
    with open(UNMATCHED_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=['row', 'cnpj_raw', 'cnpj_digits', 'reason'])
        w.writeheader()
        w.writerows(unmatched_rows)

    print('Resumo:')
    print(' - Linhas atualizadas:', atualizados)
    print(' - Sem correspondência:', sem_match)
    print('Exemplos de não casados:', exemplos_unmatched)

if __name__ == '__main__':
    main()
