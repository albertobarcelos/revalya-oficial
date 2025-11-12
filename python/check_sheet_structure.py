#!/usr/bin/env python3
"""
Script para verificar a estrutura da planilha com contract_ids
"""

import openpyxl

def main():
    """Função principal"""
    
    # Carregar Excel
    excel_path = 'python/contratos_prontos_with_ids.xlsx'
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        print("Estrutura da planilha:")
        print("="*50)
        
        # Mostrar headers
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            print(f"Coluna {col}: {header}")
        
        print("\n" + "="*50)
        print("Exemplo de dados (linha 3):")
        
        # Mostrar exemplo de dados
        for col in range(1, min(15, ws.max_column + 1)):
            value = ws.cell(row=3, column=col).value
            header = ws.cell(row=1, column=col).value
            print(f"{header}: {value}")
            
    except Exception as e:
        print(f"Erro: {e}")

if __name__ == "__main__":
    main()