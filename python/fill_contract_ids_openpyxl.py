#!/usr/bin/env python3
"""
Script para buscar contract_id pelo contract_number e preencher na planilha usando openpyxl
"""

import os
from dotenv import load_dotenv
from supabase import create_client, Client
from openpyxl import load_workbook

# Carregar variáveis de ambiente
load_dotenv()

# Configurações do Supabase
SUPABASE_URL = os.getenv('SUPABASE_URL')
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_KEY') or os.getenv('SUPABASE_KEY')
TENANT_ID = '8d2888f1-64a5-445f-84f5-2614d5160251'  # Tenant correto baseado nos dados

def get_contract_id(contract_number, supabase_client):
    """Busca o contract_id pelo contract_number"""
    try:
        response = supabase_client.table('contracts').select('id').eq('contract_number', contract_number).eq('tenant_id', TENANT_ID).execute()
        if response.data and len(response.data) > 0:
            return response.data[0]['id']
        return None
    except Exception as e:
        print(f"Erro ao buscar contract_id para {contract_number}: {e}")
        return None

def main():
    """Função principal"""
    if not SUPABASE_URL or not SUPABASE_KEY:
        print("Erro: SUPABASE_URL e SUPABASE_KEY são obrigatórios")
        return
    
    # Inicializar Supabase
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    # Caminho do arquivo
    file_path = "python/contratos_prontos.xlsx"
    
    try:
        # Carregar a planilha
        wb = load_workbook(file_path)
        ws = wb.active
        
        # Encontrar índices das colunas
        headers = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers[header] = col
        
        print("Colunas encontradas:", list(headers.keys()))
        
        # Verificar se temos as colunas necessárias
        if 'CodGE' not in headers:
            print("Erro: Coluna 'CodGE' não encontrada")
            return
        
        # Adicionar coluna contract_id se não existir
        if 'contract_id' not in headers:
            contract_id_col = ws.max_column + 1
            ws.cell(row=1, column=contract_id_col, value='contract_id')
            headers['contract_id'] = contract_id_col
            print(f"Coluna 'contract_id' adicionada na posição {contract_id_col}")
        else:
            contract_id_col = headers['contract_id']
        
        # Processar cada linha
        found_count = 0
        not_found_count = 0
        codge_col = headers['CodGE']
        
        for row in range(2, ws.max_row + 1):  # Pular header
            contract_number = ws.cell(row=row, column=codge_col).value
            
            # Pular linhas sem contract_number
            if not contract_number or str(contract_number).strip() == '':
                continue
            
            # Buscar o contract_id
            contract_id = get_contract_id(contract_number, supabase)
            
            if contract_id:
                ws.cell(row=row, column=contract_id_col, value=contract_id)
                found_count += 1
                print(f"Linha {row}: Contract {contract_number} -> ID: {contract_id}")
            else:
                not_found_count += 1
                print(f"Linha {row}: Contract {contract_number} -> NÃO ENCONTRADO")
        
        # Salvar a planilha atualizada
        output_file = "python/contratos_prontos_with_ids.xlsx"
        wb.save(output_file)
        
        print(f"\nResumo:")
        print(f"- Contratos encontrados: {found_count}")
        print(f"- Contratos não encontrados: {not_found_count}")
        print(f"- Planilha salva em: {output_file}")
        
        # Criar relatório de não encontrados
        if not_found_count > 0:
            not_found_list = []
            for row in range(2, ws.max_row + 1):
                contract_number = ws.cell(row=row, column=codge_col).value
                contract_id = ws.cell(row=row, column=contract_id_col).value
                if contract_number and not contract_id:
                    customer_id = ws.cell(row=row, column=headers.get('customer_id', 1)).value if 'customer_id' in headers else ''
                    loja = ws.cell(row=row, column=headers.get('loja', 1)).value if 'loja' in headers else ''
                    not_found_list.append({
                        'CodGE': contract_number,
                        'customer_id': customer_id,
                        'loja': loja
                    })
            
            if not_found_list:
                with open("python/contract_numbers_not_found.csv", "w") as f:
                    f.write("CodGE,customer_id,loja\n")
                    for item in not_found_list:
                        f.write(f"{item['CodGE']},{item['customer_id']},{item['loja']}\n")
                print(f"- Relatório de não encontrados salvo em: python/contract_numbers_not_found.csv")
            
    except Exception as e:
        print(f"Erro ao processar planilha: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()