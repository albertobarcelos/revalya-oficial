#!/usr/bin/env python3
"""
Script para analisar o mapeamento atual das colunas da planilha
"""

import openpyxl
from datetime import datetime

def analyze_current_mapping():
    """Analisa o mapeamento atual e identifica problemas"""
    
    print("🔍 ANALISANDO MAPEAMENTO ATUAL")
    print("="*80)
    
    # Carrega a planilha
    workbook = openpyxl.load_workbook('contratos_prontos_with_ids.xlsx')
    sheet = workbook.active
    
    # Mapeamento de colunas do script atual
    script_mapping = {
        'CONTRACT_ID_COL': 2,   # Coluna B (contract_id)
        'SERVICE_NAME_COL': 13, # Coluna M (Gestao - nome do serviço)
        'QUANTITY_COL': 12,     # Coluna L (numequipamentos)
        'UNIT_PRICE_COL': 11,   # Coluna K (Custo)
    }
    
    # Mapeamento de serviços do script
    service_columns = {
        'PDV/Comandas': 'PDV Legal',
        'NFCE': 'HIPER',
        'Estoque': 'MÓDULO DE ESTOQUE', 
        'Financeiro': 'MÓDULO FINANCEIRO',
        'Delivery Legal': 'DELIVERY LEGAL - 25K TRANSASIONADO',
        'Delivery Legal +': 'DELIVERY LEGAL - ACIMA 25K TRANSASIONADO',
        'Fidelidade legal': 'FIDELIDADE LEGAL',
        'Totem Autoatendimento': 'AUTO ATENDIMENTO BALANÇA',
        'KDS': 'KDS',
        'Balanca Auto Servico': 'AUTO ATENDIMENTO BALANÇA'
    }
    
    print("📋 MAPEAMENTO DE COLUNAS DO SCRIPT:")
    print("-" * 50)
    for key, col_num in script_mapping.items():
        header = sheet.cell(row=1, column=col_num).value
        print(f"{key}: Coluna {col_num} ({header})")
    
    print("\n🔍 ANALISANDO DADOS DA LINHA 3:")
    print("-" * 50)
    
    # Analisa a linha 3 como exemplo
    row_num = 3
    
    # Dados básicos
    contract_id = sheet.cell(row=row_num, column=2).value
    cnpj = sheet.cell(row=row_num, column=1).value
    custo = sheet.cell(row=row_num, column=11).value
    num_equip = sheet.cell(row=row_num, column=12).value
    gestao = sheet.cell(row=row_num, column=13).value
    
    print(f"Contract ID: {contract_id}")
    print(f"CNPJ: {cnpj}")
    print(f"Custo (Coluna 11): {custo} (Tipo: {type(custo)})")
    print(f"Núm Equipamentos (Coluna 12): {num_equip} (Tipo: {type(num_equip)})")
    print(f"Gestão (Coluna 13): {gestao} (Tipo: {type(gestao)})")
    
    print("\n🔄 SERVIÇOS ENCONTRADOS:")
    print("-" * 50)
    
    active_services = []
    
    # Verifica cada coluna de serviço
    for col_name, service_name in service_columns.items():
        # Procura a coluna pelo nome
        for col_num in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=1, column=col_num).value
            if header_cell and col_name in str(header_cell):
                value = sheet.cell(row=row_num, column=col_num).value
                print(f"{col_name} (Col {col_num}): {value} (Tipo: {type(value)})")
                
                # Verifica se está ativo
                is_active = False
                if value and str(value).strip().upper() in ['SIM', '1', 'YES', 'TRUE']:
                    is_active = True
                elif isinstance(value, (int, float)) and value > 0:
                    is_active = True
                
                if is_active:
                    active_services.append(service_name)
                    print(f"  ✅ ATIVO: {service_name}")
                break
    
    # Verifica Gestão
    gestao_value = sheet.cell(row=row_num, column=13).value
    if gestao_value and str(gestao_value).strip().upper() == 'SIM':
        print(f"Gestão: SIM → Adicionando HIPER GESTÃO")
        active_services.append('HIPER GESTÃO')
    
    print(f"\n📊 RESUMO DA LINHA {row_num}:")
    print("-" * 50)
    print(f"Serviços ativos encontrados: {len(active_services)}")
    print(f"Serviços: {active_services}")
    print(f"Quantidade padrão usada: 1")
    print(f"Preço unitário padrão usado: 0")
    print(f"Total calculado: {1 * 0}")
    
    print("\n⚠️  PROBLEMAS IDENTIFICADOS:")
    print("-" * 50)
    print("1. Custo (Coluna 11) está vazio/None - preço unitário ficará 0")
    print("2. Núm Equipamentos (Coluna 12) está como 1 - quantidade será 1")
    print("3. Gestão (Coluna 13) está como 'SIM' - adicionará HIPER GESTÃO")
    print("4. PDV/Comandas (Coluna 14) está como 1 - ativará PDV Legal")
    print("5. NFCE (Coluna 15) está como 'SIM' - ativará HIPER")
    
    print("\n💡 SUGESTÕES DE MELHORIA:")
    print("-" * 50)
    print("1. Usar o valor da coluna 'Custo' quando disponível")
    print("2. Usar o valor da coluna 'numequipamentos' para quantidade")
    print("3. Mapear corretamente os valores '1', 'SIM', etc.")
    print("4. Adicionar logs detalhados do mapeamento")
    print("5. Validar se o contract_id existe antes de processar")
    
    workbook.close()

if __name__ == "__main__":
    analyze_current_mapping()