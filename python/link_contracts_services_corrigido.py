#!/usr/bin/env python3
"""
Script corrigido para vincular serviços aos contratos na tabela contract_services
Versão correta: Lê IDs dos serviços da linha 2, puxa preços do banco, e usa custo específico
"""

import os
import sys
from datetime import datetime
import openpyxl
from dotenv import load_dotenv

# Importa o cliente Supabase diretamente
from supabase import create_client, Client

# Carrega variáveis de ambiente
load_dotenv()

# Configuração do tenant (será detectado dinamicamente)
DEFAULT_TENANT_ID = "8d2888f1-64a5-445f-84f5-2614d5160251"

# Service ID específico que deve usar a coluna "Custo"
COST_SERVICE_ID = "dbad5192-79b1-41e6-adbd-5218167c738c"

# Service ID que tem quantity especial (coluna N)
QUANTITY_SPECIAL_SERVICE_ID = "c1552361-c1db-43ae-ad3a-9a6f8143f668"

def get_service_info(supabase, service_id):
    """Busca informações do serviço no banco de dados"""
    try:
        result = supabase.table('services').select('name, default_price, cost_price').eq('id', service_id).execute()
        if result.data and len(result.data) > 0:
            return result.data[0]
        return None
    except Exception as e:
        print(f"❌ Erro ao buscar serviço {service_id}: {e}")
        return None

def get_contract_info(supabase, contract_id):
    """Busca informações do contrato incluindo tenant_id"""
    try:
        result = supabase.table('contracts').select('tenant_id').eq('id', contract_id).execute()
        if result.data and len(result.data) > 0:
            return result.data[0]
        return None
    except Exception as e:
        print(f"❌ Erro ao buscar contrato {contract_id}: {e}")
        return None

def get_active_services_from_row(sheet, row_num, row_data):
    """Analisa a linha e retorna lista de serviços ativos com seus IDs"""
    active_services = []
    
    print(f"\n🔍 Analisando linha {row_num}:")
    print(f"   Contract ID: {row_data.get(2, 'N/A')}")
    
    # Lê os IDs dos serviços da linha 2 (headers com IDs)
    service_mapping = {}
    for col_num in range(13, 24):  # Colunas M a X (Gestao até Balanca Auto Servico)
        service_id = sheet.cell(row=2, column=col_num).value
        if service_id and isinstance(service_id, str) and len(service_id) == 36:  # UUID válido
            # Mapeia o nome da coluna para o ID do serviço
            header_name = sheet.cell(row=1, column=col_num).value
            service_mapping[col_num] = {
                'id': service_id,
                'name': header_name,
                'col_num': col_num
            }
    
    print(f"   Serviços mapeados na linha 2: {len(service_mapping)}")
    
    # Verifica cada coluna de serviço
    for col_num, service_info in service_mapping.items():
        value = row_data.get(col_num, '')
        service_id = service_info['id']
        service_name = service_info['name']
        
        print(f"   Coluna {col_num} ({service_name}): {value}")
        
        # Verifica se o serviço está ativo
        is_active = False
        value_str = str(value).strip().upper() if value else ''
        
        # Verifica valores booleanos/texto
        if value_str in ['SIM', '1', 'YES', 'TRUE']:
            is_active = True
        elif value_str in ['NÃO', 'NAO', '0', 'NO', 'FALSE']:
            # Para valores "NÃO" ou similares, não cria vínculo
            print(f"   ❌ IGNORADO: Valor '{value}' indica que não deve criar vínculo")
            continue
        else:
            # Tenta converter para número e verificar se é positivo
            try:
                numeric_value = float(str(value).strip())
                if numeric_value > 0:
                    is_active = True
            except (ValueError, TypeError):
                # Não é um número válido, ignora
                pass
        
        if is_active:
            active_services.append({
                'id': service_id,
                'name': service_name,
                'col_num': col_num,
                'value': value
            })
            print(f"   ✅ ATIVO: {service_name} (ID: {service_id})")
    
    return active_services

def process_contract_services():
    """Processa a planilha e cria vínculos na tabela contract_services"""
    
    print("🚀 Iniciando vinculação de serviços aos contratos (VERSÃO CORRIGIDA)...")
    
    # Carrega a planilha
    workbook = openpyxl.load_workbook('contratos_prontos_with_ids.xlsx')
    sheet = workbook.active
    
    # Configura o cliente Supabase
    supabase_url = os.getenv('SUPABASE_URL')
    supabase_key = os.getenv('SUPABASE_KEY')
    
    if not supabase_url or not supabase_key:
        print("❌ Erro: Variáveis SUPABASE_URL e SUPABASE_KEY não configuradas!")
        return
    
    supabase: Client = create_client(supabase_url, supabase_key)
    
    # Contadores
    total_processado = 0
    servicos_criados = 0
    servicos_ignorados = 0
    erros = 0
    
    # Lê os dados a partir da linha 3 (pulando cabeçalho)
    # Processa todas as linhas com dados da planilha
    for row_num in range(3, sheet.max_row + 1):
        try:
            # Coleta todos os dados da linha
            row_data = {}
            for col_num in range(1, sheet.max_column + 1):
                row_data[col_num] = sheet.cell(row=row_num, column=col_num).value
            
            contract_id = row_data.get(2)  # Coluna B
            total_processado += 1
            
            # Validações
            if not contract_id:
                print(f"⚠️  Linha {row_num}: contract_id vazio, ignorando...")
                servicos_ignorados += 1
                continue
            
            # Verifica se o contrato existe e obtém tenant_id
            contract_info = get_contract_info(supabase, contract_id)
            if not contract_info:
                print(f"⚠️  Linha {row_num}: Contrato {contract_id} não encontrado, ignorando...")
                servicos_ignorados += 1
                continue
            
            tenant_id = contract_info['tenant_id']
            print(f"   Tenant ID detectado: {tenant_id}")
            
            # Obtém serviços ativos para esta linha
            active_services = get_active_services_from_row(sheet, row_num, row_data)
            
            if not active_services:
                print(f"ℹ️  Linha {row_num}: Nenhum serviço ativo encontrado, ignorando...")
                servicos_ignorados += 1
                continue
            
            print(f"   Total de serviços ativos: {len(active_services)}")
            
            # Processa cada serviço ativo
            for service in active_services:
                service_id = service['id']
                service_name = service['name']
                
                # Busca informações do serviço no banco
                service_info = get_service_info(supabase, service_id)
                if not service_info:
                    print(f"⚠️  Serviço {service_id} não encontrado no banco, ignorando...")
                    continue
                
                # Define valores base
                quantity = 1  # Padrão
                unit_price = service_info.get('default_price', 0)  # Preço do banco (coluna default_price)
                cost_price = 0  # Por padrão não usa cost_price do banco
                
                # Define quantity baseado no valor da célula
                service_value = service['value']
                
                # Tenta converter o valor para número (trata strings numéricas)
                try:
                    numeric_value = float(str(service_value).strip())
                    if numeric_value > 0:
                        quantity = int(numeric_value)
                        print(f"   📊 Quantidade numérica: {quantity}")
                except (ValueError, TypeError):
                    # Se não for número, mantém quantidade 1 para serviços ativos
                    print(f"   📊 Quantidade padrão (ativa): {quantity}")
                
                # Tratamento especial para custo - verifica coluna 11 (Custo)
                custo_value = row_data.get(11)
                if custo_value:
                    try:
                        custo_numerico = float(str(custo_value).strip())
                        if custo_numerico > 0:
                            cost_price = custo_numerico
                            print(f"   💰 Custo aplicado: R$ {cost_price}")
                    except (ValueError, TypeError):
                        pass  # Ignora se não for número válido
                
                print(f"   💵 Serviço: {service_info['name']}")
                print(f"      Quantidade: {quantity}")
                print(f"      Preço Unitário (default_price): R$ {unit_price}")
                print(f"      Custo (cost_price): R$ {cost_price}")
                print(f"      Total: R$ {quantity * unit_price}")
                
                # Verifica se já existe o vínculo
                existing = supabase.table('contract_services').select('id').eq('contract_id', contract_id).eq('service_id', service_id).execute()
                
                if existing.data:
                    print(f"ℹ️  Vínculo já existe para contrato {contract_id} e serviço {service_info['name']}, atualizando...")
                    
                    # Atualiza o vínculo existente com valores fixos conforme requisitos
                    result = supabase.table('contract_services').update({
                        'quantity': quantity,
                        'unit_price': unit_price,
                        'cost_price': cost_price,
                        'description': 'PDVLegal',  # Descrição padrão
                        'generate_billing': False,  # Deve ser FALSE
                        'billing_type': 'Único',  # Deve ser "Único"
                        'payment_method': 'Boleto',  # Deve ser "Boleto"
                        'recurrence_frequency': 'Mensal',  # Deve ser "Mensal"
                        'updated_at': datetime.now().isoformat()
                    }).eq('id', existing.data[0]['id']).execute()
                    
                    if result.data:
                        servicos_criados += 1
                        print(f"✅ Vínculo atualizado com sucesso!")
                    else:
                        erros += 1
                        print(f"❌ Erro ao atualizar vínculo")
                else:
                    # Cria novo vínculo com valores fixos conforme requisitos
                    result = supabase.table('contract_services').insert({
                        'contract_id': contract_id,
                        'service_id': service_id,
                        'quantity': quantity,
                        'unit_price': unit_price,
                        'cost_price': cost_price,
                        'description': 'PDVLegal',  # Descrição padrão
                        'tenant_id': tenant_id,
                        'is_active': True,
                        'no_charge': False,
                        'generate_billing': False,  # Deve ser FALSE
                        'billing_type': 'Único',  # Deve ser "Único"
                        'payment_method': 'Boleto',  # Deve ser "Boleto"
                        'recurrence_frequency': 'Mensal',  # Deve ser "Mensal"
                        'due_type': 'days_after_billing',
                        'due_value': 5,
                        'installments': 1
                    }).execute()
                    
                    if result.data:
                        servicos_criados += 1
                        print(f"✅ Vínculo criado com sucesso!")
                    else:
                        erros += 1
                        print(f"❌ Erro ao criar vínculo")
            
            # Progresso a cada 10 linhas
            if total_processado % 10 == 0:
                print(f"\n📊 Progresso: {total_processado} linhas processadas...")
                print(f"   Vínculos criados/atualizados: {servicos_criados}")
                print(f"   Erros: {erros}")
                print(f"   Ignorados: {servicos_ignorados}")
                print("-" * 60)
                
        except Exception as e:
            erros += 1
            print(f"❌ Linha {row_num}: Erro inesperado: {str(e)}")
            continue
    
    # Relatório final
    print("\n" + "="*80)
    print("📋 RELATÓRIO FINAL")
    print("="*80)
    print(f"Total de linhas processadas: {total_processado}")
    print(f"Vínculos criados/atualizados: {servicos_criados}")
    print(f"Linhas ignoradas: {servicos_ignorados}")
    print(f"Erros: {erros}")
    print("="*80)
    
    workbook.close()
    return servicos_criados

def main():
    """Função principal"""
    try:
        print("🔍 Verificando arquivos necessários...")
        
        # Verifica se a planilha existe
        if not os.path.exists('contratos_prontos_with_ids.xlsx'):
            print("❌ Erro: Planilha 'contratos_prontos_with_ids.xlsx' não encontrada!")
            return
        
        print("✅ Planilha encontrada!")
        
        # Executa o processamento
        servicos_criados = process_contract_services()
        
        if servicos_criados > 0:
            print(f"\n🎉 Sucesso! {servicos_criados} vínculos criados/atualizados.")
        else:
            print("\n⚠️  Nenhum vínculo foi criado. Verifique os logs acima.")
            
    except Exception as e:
        print(f"❌ Erro fatal: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()