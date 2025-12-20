import os
import re
import csv
import unicodedata
import datetime
from typing import Optional, Dict, Any, List
from openpyxl import load_workbook
from dotenv import load_dotenv

# Supabase
from supabase import create_client, Client

# Diretório base relativo ao arquivo atual
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_XLSX = os.path.join(BASE_DIR, 'contratos.xlsx')
OUTPUT_XLSX = os.path.join(BASE_DIR, 'contratos_com_ids.xlsx')
UNMATCHED_CSV = os.path.join(BASE_DIR, 'contratos_ids_unmatched.csv')

def normalize_name(name: str) -> str:
    if not isinstance(name, str):
        name = str(name)
    nf = unicodedata.normalize('NFD', name)
    s = ''.join(c for c in nf if unicodedata.category(c) != 'Mn')
    return re.sub(r'\s+', '', s).lower()

def find_sheet(wb, targets: List[str]):
    """Prefere a ordem dos alvos fornecidos. Se não encontrar, cai na primeira planilha."""
    targets_norm = [normalize_name(t) for t in targets]
    for tn in targets_norm:
        for nm in wb.sheetnames:
            if normalize_name(nm) == tn:
                return wb[nm]
    # fallback: primeira planilha
    return wb[wb.sheetnames[0]]

DIGITS_RE = re.compile(r'\D+')

def only_digits(value: Optional[object]) -> str:
    if value is None:
        return ''
    s = str(value).strip()
    return DIGITS_RE.sub('', s)

def strip_leading_zeros(s: str) -> str:
    return s.lstrip('0') or '0'

def get_supabase_client() -> Client:
    load_dotenv()  # carrega .env
    url = os.getenv('SUPABASE_URL')
    key = os.getenv('SUPABASE_KEY')
    if not url or not key:
        raise RuntimeError('SUPABASE_URL e SUPABASE_KEY não encontrados no ambiente. Configure o arquivo python/.env.')
    return create_client(url, key)

def fetch_customers_map(client: Client, tenant_id: Optional[str] = None) -> Dict[str, Dict[str, Any]]:
    """
    Retorna um mapa de documento -> customer ({id, tenant_id, name, company}).
    Documento é cpf_cnpj vindo como texto (sem zeros à esquerda, pois está salvo como bigint no banco).
    """
    start = 0
    step = 1000
    mapping: Dict[str, Dict[str, Any]] = {}
    duplicates: Dict[str, int] = {}
    while True:
        query = client.table('customers').select('id,tenant_id,name,company,cpf_cnpj').range(start, start + step - 1)
        if tenant_id:
            query = query.eq('tenant_id', tenant_id)
        resp = query.execute()
        rows = resp.data or []
        for row in rows:
            doc = row.get('cpf_cnpj')
            if doc is None:
                continue
            doc_str = str(doc).strip()
            # chave será sem zeros à esquerda, pois bigint perde zeros naturais
            key = strip_leading_zeros(only_digits(doc_str))
            if key:
                if key in mapping:
                    duplicates[key] = duplicates.get(key, 1) + 1
                mapping[key] = {
                    'id': row.get('id'),
                    'tenant_id': row.get('tenant_id'),
                    'name': row.get('name'),
                    'company': row.get('company'),
                    'raw': doc_str,
                }
        if len(rows) < step:
            break
        start += step
    if duplicates:
        print('Atenção: Foram detectadas chaves duplicadas em cpf_cnpj (possíveis clientes repetidos em tenants diferentes):', len(duplicates))
        # imprime alguns exemplos
        shown = 0
        for k, cnt in duplicates.items():
            print(f' - Documento {k} aparece {cnt} vezes (considere filtrar por TENANT_ID)')
            shown += 1
            if shown >= 10:
                break
    return mapping

def detect_document_column(sheet) -> int:
    """Procura por coluna com cabeçalho contendo 'cnpj' ou 'cpf' ou 'documento'."""
    header_row = 1
    for c in range(1, sheet.max_column + 1):
        hv = sheet.cell(row=header_row, column=c).value
        if isinstance(hv, str):
            lower = hv.lower()
            if 'cnpj' in lower or 'cpf' in lower or 'document' in lower or 'documento' in lower:
                return c
    return 1  # fallback para coluna A

def main():
    print('Abrindo arquivo:', INPUT_XLSX)
    wb = load_workbook(INPUT_XLSX, data_only=True)
    sheet = find_sheet(wb, ['contratos', 'LicencasAtivas'])
    print('Planilha selecionada:', sheet.title)

    doc_col = detect_document_column(sheet)
    print('Coluna de documento (CPF/CNPJ) detectada:', doc_col)

    # Criar coluna de destino customer_id
    dest_header = os.getenv('DEST_HEADER', 'customer_id')
    # Descobrir se já existe
    dest_col = None
    for c in range(1, sheet.max_column + 1):
        hv = sheet.cell(row=1, column=c).value
        if hv == dest_header:
            dest_col = c
            break
    if dest_col is None:
        dest_col = sheet.max_column + 1
        sheet.cell(row=1, column=dest_col).value = dest_header
    print('Coluna de destino:', dest_col, 'Header:', dest_header)

    # Supabase
    client = get_supabase_client()
    tenant_id = os.getenv('TENANT_ID')  # opcional
    mapping = fetch_customers_map(client, tenant_id=tenant_id)
    print('Clientes carregados do Supabase:', len(mapping))

    atualizados = 0
    sem_match = 0
    unmatched_rows = []
    exemplos_unmatched = []

    for r in range(2, sheet.max_row + 1):
        raw = sheet.cell(row=r, column=doc_col).value
        digits = only_digits(raw)
        digits_stripped = strip_leading_zeros(digits)
        if not digits_stripped or digits_stripped == '0':
            sem_match += 1
            unmatched_rows.append({
                'row': r,
                'document_raw': '' if raw is None else str(raw),
                'document_digits': digits,
                'key_used': digits_stripped,
                'reason': 'Documento vazio'
            })
            if len(exemplos_unmatched) < 10:
                exemplos_unmatched.append((r, raw, digits, 'Documento vazio'))
            continue
        cust = mapping.get(digits_stripped)
        if cust:
            sheet.cell(row=r, column=dest_col).value = cust['id']
            atualizados += 1
        else:
            sem_match += 1
            unmatched_rows.append({
                'row': r,
                'document_raw': '' if raw is None else str(raw),
                'document_digits': digits,
                'key_used': digits_stripped,
                'reason': 'sem correspondência'
            })
            if len(exemplos_unmatched) < 10:
                exemplos_unmatched.append((r, raw, digits_stripped, 'sem correspondência'))

    print('Salvando arquivo atualizado em:', OUTPUT_XLSX)
    try:
        wb.save(OUTPUT_XLSX)
    except PermissionError:
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        alt_output = os.path.join(BASE_DIR, f'contratos_com_ids_{ts}.xlsx')
        print('Arquivo de saída em uso ou sem permissão. Salvando como:', alt_output)
        wb.save(alt_output)

    print('Gerando relatório de não casados em:', UNMATCHED_CSV)
    with open(UNMATCHED_CSV, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=['row', 'document_raw', 'document_digits', 'key_used', 'reason'])
        w.writeheader()
        w.writerows(unmatched_rows)

    print('Resumo:')
    print(' - Linhas atualizadas:', atualizados)
    print(' - Sem correspondência:', sem_match)
    print('Exemplos de não casados:', exemplos_unmatched)

if __name__ == '__main__':
    main()