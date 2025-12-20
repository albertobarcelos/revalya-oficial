import os
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client, Client


TENANT_ID_DEFAULT = "8d2888f1-64a5-445f-84f5-2614d5160251"
CONTRACTS_TABLE = "contracts"


def _strip_accents(text: str) -> str:
    import unicodedata
    return (
        unicodedata.normalize("NFKD", text)
        .encode("ASCII", "ignore")
        .decode("ASCII")
    )


def _norm_header(s: Optional[str]) -> str:
    if s is None:
        return ""
    s2 = _strip_accents(str(s))
    return (
        s2.strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("/", "_")
    )


def _parse_date(value) -> Optional[str]:
    from datetime import datetime
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if isinstance(value, datetime):
        try:
            return value.date().isoformat()
        except Exception:
            pass
    # tentar dd/mm/yyyy
    try:
        return datetime.strptime(str(value), "%d/%m/%Y").date().isoformat()
    except Exception:
        # tentar ISO
        try:
            return datetime.fromisoformat(str(value)).date().isoformat()
        except Exception:
            return None


def _parse_int(value) -> Optional[int]:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    try:
        return int(str(value).strip())
    except Exception:
        return None


def _parse_number(value) -> Optional[float]:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    try:
        return float(str(value).replace(".", "").replace(",", "."))
    except Exception:
        return None


def _get_client() -> Client:
    load_dotenv()
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_KEY") or os.environ.get("SUPABASE_KEY")
    if not url or not key:
        raise RuntimeError("Defina SUPABASE_URL e SUPABASE_SERVICE_KEY/SUPABASE_KEY em python/.env")
    return create_client(url, key)


def _build_row_mapper(headers: List[str]) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    for i, h in enumerate(headers, start=1):
        idx[_norm_header(h)] = i
    return idx


def read_rows(xlsx_path: Path) -> Tuple[List[Dict[str, Any]], List[str]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    idx = _build_row_mapper(headers)

    # Mapeamento de colunas do Excel para o banco de dados
    synonyms: Dict[str, List[str]] = {
        # IDs e números
        "customer_id": ["customer_id", "id_cliente", "cliente_id"],
        "codge": ["codge", "codigo", "numero_contrato", "contract_number", "CodGE"],
        
        # Dados do cliente
        "cnpj": ["cnpj", "cpf", "cpf_cnpj", "documento"],
        "grupo_economico": ["grupoeconomico", "grupo_economico", "Grupoeconomico", "Grupo Economico"],
        "loja": ["loja", "filial", "unidade", "estabelecimento"],
        "email": ["email", "e_mail", "correio_eletronico"],
        
        # Datas
        "data_inicio": ["data_inicio", "data_inicial", "inicio", "initial_date", "Ativacao"],
        "data_fim": ["data_fim", "data_final", "fim", "final_date"],
        
        # Valores e configurações
        "valor_total": ["valor_total", "valor", "total", "total_amount", "Custo", "custo"],
        "tipo_faturamento": ["tipo_faturamento", "faturamento", "billing_type", "TipoNegocioDetalhes", "tipo_negocio"],
        "dia_faturamento": ["dia_faturamento", "dia_vencimento", "billing_day"],
        "status": ["status", "situacao", "estado"],
        "descricao": ["descricao", "descricao_contrato", "description", "observacoes"],
        
        # Quantidades
        "num_equipamentos": ["numequipamentos", "numero_equipamentos", "quantidade_equipamentos", "num_equipamentos"],
    }

    # normaliza sinônimos
    synonyms = {k: [_norm_header(v) for v in vs] for k, vs in synonyms.items()}

    def col_index_for(key: str) -> Optional[int]:
        # procura pelo primeiro sinônimo presente
        for candidate in synonyms.get(key, [key]):
            if candidate in idx:
                return idx[candidate]
        return None

    required = [
        "customer_id",
        "codge",  # contract_number
        "data_inicio",
        "data_fim",
        "tipo_faturamento",
        "dia_faturamento",
    ]
    erros: List[str] = []
    for r in required:
        if col_index_for(r) is None:
            erros.append(f"Cabeçalho obrigatório ausente: {r}")

    registros: List[Dict[str, Any]] = []

    for row in range(2, ws.max_row + 1):
        # parar em linhas totalmente vazias
        row_vals = [ws.cell(row=row, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_vals):
            continue

        def get(name: str):
            col = col_index_for(name)
            return ws.cell(row=row, column=col).value if col else None

        contract_number = get("codge")
        customer_id = get("customer_id")
        initial_date = _parse_date(get("data_inicio"))
        final_date = _parse_date(get("data_fim"))
        billing_type = get("tipo_faturamento")
        billing_day = _parse_int(get("dia_faturamento"))

        # opcionais
        total_amount = None
        for guess in ("valor_total", "valor_contrato", "total_amount", "valor"):
            v = get(guess)
            if v is not None and str(v).strip() != "":
                total_amount = _parse_number(v)
                break
        description = None
        for guess in ("descricao", "observacoes", "description", "observacoes_do_contrato"):
            v = get(guess)
            if v is not None and str(v).strip() != "":
                description = str(v)
                break
        status = None
        vstatus = get("status")
        if vstatus is not None and str(vstatus).strip() != "":
            status = str(vstatus).strip().upper()

        # validações mínimas
        missing = []
        if not customer_id:
            missing.append("customer_id")
        if not contract_number:
            missing.append("codge")
        if not initial_date:
            missing.append("data_inicio")

        if missing:
            erros.append(f"Linha {row}: faltam campos obrigatórios: {', '.join(missing)}")
            continue

        # valores padrão para campos opcionais
        if not final_date:
            from datetime import datetime, timedelta
            try:
                initial = datetime.fromisoformat(initial_date)
                final_date = (initial + timedelta(days=365)).date().isoformat()
            except Exception:
                final_date = None
        if not billing_type:
            billing_type = "mensal"
        if billing_day is None:
            billing_day = 1
        # garantir total_amount = 0
        total_amount = 0

        registros.append(
            {
                "tenant_id": TENANT_ID_DEFAULT,
                "customer_id": customer_id,
                "contract_number": str(contract_number),
                "status": status or "DRAFT",
                "initial_date": initial_date,
                "final_date": final_date,
                "billing_type": str(billing_type),
                "billing_day": billing_day,
                "anticipate_weekends": True,
                "reference_period": None,
                "installments": 1,
                "total_amount": total_amount or 0,
                "total_discount": 0,
                "total_tax": 0,
                "stage_id": None,
                "description": description,
                "internal_notes": "IMPORTADO POR PLANILHA",
                "billed": False,
            }
        )

    return registros, erros


def upsert_contracts(client: Client, registros: List[Dict[str, Any]]) -> Tuple[int, int, List[str]]:
    """Tenta upsert; se não suportado, faz insert com verificação prévia de existência."""
    errors: List[str] = []

    # Tenta upsert em lote
    try:
        resp = client.table(CONTRACTS_TABLE).upsert(registros).execute()
        # Não dá para distinguir update/insert facilmente aqui
        inserted = len(resp.data or [])
        return inserted, 0, errors
    except Exception as e:
        errors.append(f"Upsert não suportado ou falhou: {e}")

    # Fallback: insert somente se não existir (tenant_id, contract_number)
    inserted = 0
    updated = 0
    for r in registros:
        try:
            exists = (
                client
                .table(CONTRACTS_TABLE)
                .select("id")
                .eq("tenant_id", r["tenant_id"])
                .eq("contract_number", r["contract_number"])
                .limit(1)
                .execute()
            )
            if exists.data:
                # Atualiza campos relevantes
                upd = client.table(CONTRACTS_TABLE).update({
                    "customer_id": r["customer_id"],
                    "status": r["status"],
                    "initial_date": r["initial_date"],
                    "final_date": r["final_date"],
                    "billing_type": r["billing_type"],
                    "billing_day": r["billing_day"],
                    "anticipate_weekends": r["anticipate_weekends"],
                    "reference_period": r["reference_period"],
                    "installments": r["installments"],
                    "total_amount": r["total_amount"],
                    "total_discount": r["total_discount"],
                    "total_tax": r["total_tax"],
                    "stage_id": r["stage_id"],
                    "description": r["description"],
                    "internal_notes": r["internal_notes"],
                    "billed": r["billed"],
                }).eq("tenant_id", r["tenant_id"]).eq("contract_number", r["contract_number"]).execute()
                updated += len(upd.data or [])
            else:
                ins = client.table(CONTRACTS_TABLE).insert(r).execute()
                inserted += len(ins.data or [])
        except Exception as e:
            errors.append(f"Falha ao inserir/atualizar contrato {r['contract_number']}: {e}")

    return inserted, updated, errors


def main():
    # caminho padrão relativo ao repo
    base = Path(__file__).resolve().parent
    default_xlsx = base.parent / "python" / "contratos_prontos.xlsx"
    xlsx_path_str = os.environ.get("CONTRATOS_XLSX", str(default_xlsx))
    xlsx_path = Path(xlsx_path_str)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {xlsx_path}")

    registros, erros = read_rows(xlsx_path)
    print(f"Registros válidos: {len(registros)}")
    if erros:
        print("Erros de validação:")
        for e in erros:
            print(f" - {e}")
        # Se houver muitos erros, aborta para evitar gravações inválidas
        # Comente se desejar importar mesmo assim
        # raise SystemExit(1)

    client = _get_client()
    inserted, updated, write_errors = upsert_contracts(client, registros)
    print("Resumo de importação:")
    print(f" - Inseridos: {inserted}")
    print(f" - Atualizados: {updated}")
    if write_errors:
        print("Erros de escrita:")
        for e in write_errors:
            print(f" - {e}")


if __name__ == "__main__":
    main()