#!/usr/bin/env python3
"""
Script para vincular serviços aos contratos na tabela contract_services
Lê a planilha contratos_prontos_with_ids.xlsx e cria os vínculos
"""

import os
import sys
from datetime import datetime
import openpyxl
from dotenv import load_dotenv

# Importa o cliente Supabase diretamente
from supabase import create_client, Client

# Carrega variáveis de ambiente
load_dotenv()

# Configuração do tenant
TENANT_ID = "8d2888f1-64a5-445f-84f5-2614d5160251"

# Mapeamento de serviços da planilha para IDs do Supabase
SERVICE_MAPPING = {
    'PDV Legal': 'b8be3fd6-82f9-467a-8673-6fd12e23ff9b',
    'PDV LEGAL FULL': '4ba2ade6-1d02-4b93-b854-f12e4e719e98',
    'HIPER': '1d27cd41-434a-4fe1-ada3-eee1e3caa16e',
    'HIPER CONNECTION': '59d94ea7-0ccf-4eda-9e9d-0578883721df',
    'HIPER TEF': '437ba603-f2e7-4f4e-a102-3efd1ec6afba',
    'HIPER - NFS-e': '468694dc-b536-43b8-a588-45f4f1d19cfe',
    'HIPER - MDF-e': '63266213-e0ae-450b-ac62-84ac87a1f93b',
    'HIPER - USUARIO ADICIONAL': 'c5dda423-e3d5-4bb3-be4b-af65b389c2c6',
    'HIPER - LOJA ADICIONAL': 'de84047e-0dbe-490b-9130-193e765df8fe',
    'HIPER - CAIXA ADICIONAL': 'cb811b7b-010d-49c1-8ce6-f54f9a4c2f1f',
    'HIPER - BOLETOS BANCARIOS': '340e3787-3a56-415f-94f3-6b151add59d0',
    'HIPER - ARQUIVOS FISCAIS': '72e72da2-0a49-413c-842d-3ad03a9369ef',
    'HIPER - VAREJO DIGITAL': '256692bf-0137-4046-995a-a2d6ee72716b',
    'HIPER - IMAGEM PRODUTOS': 'fca955e2-23a8-49ef-9b52-3a71fced18fc',
    'HIPER GESTÃO': '16e7d726-27b1-4239-b82e-8209a634e2b4',
    'HIPER VENDAS': 'df1469ad-36ae-433a-90f6-2db7d1a82fc2',
    'MÓDULO FINANCEIRO': '2c343d48-fea9-4002-9106-4a324b5a5189',
    'MÓDULO FISCAL': 'c8cb99e1-3cea-4a99-ae93-5de95d45e39f',
    'MÓDULO DE ESTOQUE': '86f31600-69f9-4426-82f4-ff9f92c54021',
    'AUTO ATENDIMENTO BALANÇA': 'f02b94b3-ce18-47d2-8d6c-9b989f7fb5a5',
    'AUTO ATENDIMENTO BALANÇA ADICIONAL': 'fb2a30df-a75c-47eb-97fd-0627cc1999a0',
    'AUTO ATENDIMENTO FOOD': '565022c0-a844-412a-a81b-e41529c513c3',
    'AUTO ATENDIMENTO - VALIDAR CPF': 'cdfcc30b-ae6b-489c-a80b-3fa3cbbd65df',
    'KDS': 'c3e1864d-035b-4e12-8144-95ad7932c7da',
    'DELIVERY LEGAL - 25K TRANSASIONADO': '8b009d88-9219-4e97-90a2-8bb3677b8ec7',
    'DELIVERY LEGAL - ACIMA 25K TRANSASIONADO': '6f215517-b962-4544-a6b9-111555809e14',
    'FIDELIDADE LEGAL': '3e64cc33-9948-47d2-8d13-9cdb757c8bf1',
    'SERVIDOR EM NUVEM': '19a6b2c9-9900-4501-92de-ae72adee79a3',
    'CLOUD - RETAGUARDA': 'dbad5192-79b1-41e6-adbd-5218167c738c',
    'BI LEGAL - 10 LOJAS': 'ea227a87-1e94-4e64-abf8-be63e4833117',
    'BI LEGAL - ACIMA DE 10 LOJAS': '0bd6c314-ec17-40b2-b125-17a0d75bab7b',
    'CERTIFICADO DIGITAL': '03c568af-4047-4d55-a1bb-25cfeb8d6e12',
    'TERMINAL': 'c1552361-c1db-43ae-ad3a-9a6f8143f668',
    'IMPLANTAÇÃO PDVLEGAL': '0a68d92c-78ce-4a4b-b207-cf8b75e55bbd',
    'Jaturat': 'a0c91618-45bf-4625-ab39-eea5d4a0f54f'
}

def get_active_services(row_data, sheet):
    """Analisa a linha e retorna lista de serviços ativos"""
    active_services = []
    
    # Mapeamento de colunas para serviços
    service_columns = {
        'PDV/Comandas': 'PDV Legal',
        'NFCE': 'HIPER',
        'Estoque': 'MÓDULO DE ESTOQUE', 
        'Financeiro': 'MÓDULO FINANCEIRO',
        'Delivery Legal': 'DELIVERY LEGAL - 25K TRANSASIONADO',
        'Delivery Legal +': 'DELIVERY LEGAL - ACIMA 25K TRANSASIONADO',
        'Fidelidade legal': 'FIDELIDADE LEGAL',
        'Totem Autoatendimento': 'AUTO ATENDIMENTO BALANÇA',
        'KDS': 'KDS',
        'Balanca Auto Servico': 'AUTO ATENDIMENTO BALANÇA'
    }
    
    # Verifica cada coluna de serviço
    for col_name, service_name in service_columns.items():
        # Procura a coluna pelo nome
        for col_num in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=2, column=col_num).value  # Linha 2 tem os headers
            if header_cell and col_name in str(header_cell):
                value = row_data.get(col_num, 0)
                if value and str(value).strip().upper() in ['SIM', '1', 'YES', 'TRUE']:
                    active_services.append(service_name)
                elif isinstance(value, (int, float)) and value > 0:
                    active_services.append(service_name)
                break
    
    # Verifica Gestao (se for SIM, adiciona HIPER GESTÃO)
    gestao_value = row_data.get(13, '')  # Coluna M
    if gestao_value and str(gestao_value).strip().upper() == 'SIM':
        active_services.append('HIPER GESTÃO')
    
    return list(set(active_services))  # Remove duplicatas

def process_contract_services():
    """Processa a planilha e cria vínculos na tabela contract_services"""
    
    print("🚀 Iniciando vinculação de serviços aos contratos...")
    
    # Carrega a planilha
    workbook = openpyxl.load_workbook('python/contratos_prontos_with_ids.xlsx')
    sheet = workbook.active
    
    # Configura o cliente Supabase
    supabase_url = os.getenv('SUPABASE_URL')
    supabase_key = os.getenv('SUPABASE_KEY')
    
    if not supabase_url or not supabase_key:
        print("❌ Erro: Variáveis SUPABASE_URL e SUPABASE_KEY não configuradas!")
        return
    
    supabase: Client = create_client(supabase_url, supabase_key)
    
    # Contadores
    total_processado = 0
    servicos_criados = 0
    servicos_ignorados = 0
    erros = 0
    
    # Mapeamento de colunas (atualizado com base na nova estrutura)
    CONTRACT_ID_COL = 2   # Coluna B (contract_id)
    SERVICE_NAME_COL = 13 # Coluna M (Gestao - nome do serviço)
    QUANTITY_COL = 12     # Coluna L (numequipamentos)
    UNIT_PRICE_COL = 11   # Coluna K (Custo)
    
    # Lê os dados a partir da linha 3 (pulando cabeçalho)
    # Limita até a linha 238 onde terminam os dados válidos
    max_row_with_data = 238
    for row_num in range(3, min(sheet.max_row + 1, max_row_with_data + 1)):
        try:
            contract_id = sheet.cell(row=row_num, column=CONTRACT_ID_COL).value
            
            # Coleta todos os dados da linha
            row_data = {}
            for col_num in range(1, sheet.max_column + 1):
                row_data[col_num] = sheet.cell(row=row_num, column=col_num).value
            
            total_processado += 1
            
            # Validações
            if not contract_id:
                print(f"⚠️  Linha {row_num}: contract_id vazio, ignorando...")
                servicos_ignorados += 1
                continue
            
            # Obtém serviços ativos para esta linha
            active_services = get_active_services(row_data, sheet)
            
            if not active_services:
                print(f"ℹ️  Linha {row_num}: Nenhum serviço ativo encontrado, ignorando...")
                servicos_ignorados += 1
                continue
            
            # Processa cada serviço ativo
            for service_name in active_services:
                service_id = SERVICE_MAPPING.get(service_name)
                if not service_id:
                    print(f"⚠️  Linha {row_num}: Serviço '{service_name}' não encontrado no mapeamento, ignorando...")
                    continue
                
                # Obtém quantidade e preço (padrões)
                quantity = 1  # Padrão
                unit_price = 0  # Padrão
                
                # Tenta obter valores específicos do serviço
                if service_name == 'PDV Legal':
                    pdv_value = row_data.get(14, 0)  # Coluna N
                    if pdv_value and str(pdv_value).isdigit():
                        quantity = int(pdv_value)
                
                # Calcula valores (total_amount é coluna gerada, não precisa ser inserida)
                total_amount = quantity * unit_price
                
                # Verifica se já existe o vínculo
                existing = supabase.table('contract_services').select('id').eq('contract_id', contract_id).eq('service_id', service_id).execute()
                
                if existing.data:
                    print(f"ℹ️  Linha {row_num}: Vínculo já existe para contrato {contract_id} e serviço {service_name}, atualizando...")
                    
                    # Atualiza o vínculo existente
                    result = supabase.table('contract_services').update({
                        'quantity': quantity,
                        'unit_price': unit_price,
                        'updated_at': datetime.now().isoformat()
                    }).eq('id', existing.data[0]['id']).execute()
                    
                    if result.data:
                        servicos_criados += 1
                        print(f"✅ Linha {row_num}: Vínculo atualizado com sucesso!")
                    else:
                        erros += 1
                        print(f"❌ Linha {row_num}: Erro ao atualizar vínculo")
                else:
                    # Cria novo vínculo
                    result = supabase.table('contract_services').insert({
                        'contract_id': contract_id,
                        'service_id': service_id,
                        'quantity': quantity,
                        'unit_price': unit_price,
                        'tenant_id': TENANT_ID,
                        'is_active': True,
                        'no_charge': False,
                        'generate_billing': True,
                        'due_type': 'days_after_billing',
                        'due_value': 5,
                        'installments': 1
                    }).execute()
                    
                    if result.data:
                        servicos_criados += 1
                        print(f"✅ Linha {row_num}: Vínculo criado com sucesso!")
                    else:
                        erros += 1
                        print(f"❌ Linha {row_num}: Erro ao criar vínculo")
            
            # Progresso a cada 50 linhas
            if total_processado % 50 == 0:
                print(f"📊 Progresso: {total_processado} linhas processadas...")
                
        except Exception as e:
            erros += 1
            if 'row_num' in locals():
                print(f"❌ Linha {row_num}: Erro inesperado: {str(e)}")
            else:
                print(f"❌ Erro inesperado na inicialização: {str(e)}")
            continue
    
    # Relatório final
    print("\n" + "="*60)
    print("📋 RELATÓRIO FINAL")
    print("="*60)
    print(f"Total de linhas processadas: {total_processado}")
    print(f"Vínculos criados/atualizados: {servicos_criados}")
    print(f"Linhas ignoradas: {servicos_ignorados}")
    print(f"Erros: {erros}")
    print("="*60)
    
    workbook.close()
    return servicos_criados

def main():
    """Função principal"""
    try:
        print("🔍 Verificando arquivos necessários...")
        
        # Verifica se a planilha existe
        if not os.path.exists('python/contratos_prontos_with_ids.xlsx'):
            print("❌ Erro: Planilha 'contratos_prontos_with_ids.xlsx' não encontrada!")
            return
        
        print("✅ Planilha encontrada!")
        
        # Executa o processamento
        servicos_criados = process_contract_services()
        
        if servicos_criados > 0:
            print(f"\n🎉 Sucesso! {servicos_criados} vínculos criados/atualizados.")
        else:
            print("\n⚠️  Nenhum vínculo foi criado. Verifique os logs acima.")
            
    except Exception as e:
        print(f"❌ Erro fatal: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    main()